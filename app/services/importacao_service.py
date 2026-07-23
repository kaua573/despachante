"""
Importação e exportação em massa de clientes (com seus veículos) via Excel.

Reaproveita ClienteService/VeiculoService para criar/atualizar registros —
ou seja, passa pelas mesmas validações e regras de duplicidade (CPF, placa)
que já valem para o cadastro manual, sem duplicar essa lógica aqui.
"""
from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.cliente import Cliente
from app.services.cliente_service import ClienteService
from app.services.veiculo_service import VeiculoService
from app.services.validacao_service import (
    validar_campos_cliente, validar_campos_veiculo,
    normalizar_cpf, normalizar_telefone,
)

COLUNAS_CLIENTES = ["Nome*", "CPF*", "Telefone*", "Email", "Observação"]
COLUNAS_VEICULOS = [
    "CPF do Cliente*", "Placa*", "RENAVAM*", "Proprietário*",
    "Marca/Modelo*", "Situação", "Espécie", "Observação",
]

SITUACOES_VALIDAS = {"ativo", "desativado", "vendido"}
ESPECIES_VALIDAS = {"passeio", "carga", "reboque"}


class ImportacaoService:
    def __init__(self, session: Session, upload_dir: str = "") -> None:
        self._session = session
        self._cliente_svc = ClienteService(session, upload_dir)
        self._veiculo_svc = VeiculoService(session)

    # ── Exportação ───────────────────────────────────────────────────────────

    def exportar(self) -> bytes:
        """Gera um .xlsx com duas abas: Clientes e Veículos (ligados pelo CPF)."""
        wb = Workbook()

        fill_header = PatternFill(start_color="1A4F8A", end_color="1A4F8A", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)

        ws_clientes = wb.active
        ws_clientes.title = "Clientes"
        self._escrever_cabecalho(ws_clientes, COLUNAS_CLIENTES, fill_header, font_header)

        clientes = self._session.query(Cliente).order_by(Cliente.nome).all()
        for row_idx, c in enumerate(clientes, start=2):
            ws_clientes.cell(row=row_idx, column=1, value=c.nome)
            ws_clientes.cell(row=row_idx, column=2, value=c.cpf or "")
            ws_clientes.cell(row=row_idx, column=3, value=c.telefone or "")
            ws_clientes.cell(row=row_idx, column=4, value=c.email or "")
            ws_clientes.cell(row=row_idx, column=5, value=c.observacao or "")

        ws_veiculos = wb.create_sheet("Veiculos")
        self._escrever_cabecalho(ws_veiculos, COLUNAS_VEICULOS, fill_header, font_header)

        row_idx = 2
        for c in clientes:
            for v in sorted(c.veiculos, key=lambda x: x.placa):
                ws_veiculos.cell(row=row_idx, column=1, value=c.cpf or "")
                ws_veiculos.cell(row=row_idx, column=2, value=v.placa)
                ws_veiculos.cell(row=row_idx, column=3, value=v.renavam or "")
                ws_veiculos.cell(row=row_idx, column=4, value=v.proprietario or "")
                ws_veiculos.cell(row=row_idx, column=5, value=v.marca_modelo or "")
                ws_veiculos.cell(row=row_idx, column=6, value=v.situacao or "ativo")
                ws_veiculos.cell(row=row_idx, column=7, value=v.especie or "passeio")
                ws_veiculos.cell(row=row_idx, column=8, value=v.observacao or "")
                row_idx += 1

        for ws, colunas in ((ws_clientes, COLUNAS_CLIENTES), (ws_veiculos, COLUNAS_VEICULOS)):
            for col_idx, titulo in enumerate(colunas, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = max(14, len(titulo) + 4)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _escrever_cabecalho(ws, colunas, fill, font) -> None:
        for col_idx, titulo in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.fill = fill
            cell.font = font

    def gerar_modelo(self) -> bytes:
        """Planilha vazia, só com os cabeçalhos — para quem for preencher do zero."""
        wb = Workbook()
        fill_header = PatternFill(start_color="1A4F8A", end_color="1A4F8A", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)

        ws_clientes = wb.active
        ws_clientes.title = "Clientes"
        self._escrever_cabecalho(ws_clientes, COLUNAS_CLIENTES, fill_header, font_header)

        ws_veiculos = wb.create_sheet("Veiculos")
        self._escrever_cabecalho(ws_veiculos, COLUNAS_VEICULOS, fill_header, font_header)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── Importação ───────────────────────────────────────────────────────────

    def importar(self, conteudo_arquivo: bytes) -> dict:
        """
        Lê um .xlsx no mesmo formato do exportar() e cria/atualiza clientes e
        veículos. Cliente é identificado pelo CPF; veículo, pela placa dentro
        do mesmo cliente. Continua processando mesmo se uma linha falhar —
        cada erro é reportado, sem travar o restante do arquivo.
        """
        resultado = {
            "clientes_criados": 0, "clientes_atualizados": 0,
            "veiculos_criados": 0, "veiculos_atualizados": 0,
            "erros": [],
        }

        try:
            wb = load_workbook(io.BytesIO(conteudo_arquivo), read_only=True, data_only=True)
        except Exception:
            resultado["erros"].append({"linha": "-", "erro": "Arquivo inválido. Envie um .xlsx no formato do modelo."})
            return resultado

        if "Clientes" not in wb.sheetnames:
            resultado["erros"].append({"linha": "-", "erro": 'Aba "Clientes" não encontrada na planilha.'})
            return resultado

        cpf_para_cliente_id: dict[str, int] = {}
        self._importar_clientes(wb["Clientes"], resultado, cpf_para_cliente_id)

        if "Veiculos" in wb.sheetnames:
            self._importar_veiculos(wb["Veiculos"], resultado, cpf_para_cliente_id)

        return resultado

    def _importar_clientes(self, ws, resultado: dict, cpf_para_cliente_id: dict) -> None:
        linhas = ws.iter_rows(min_row=2, values_only=True)
        for numero_linha, linha in enumerate(linhas, start=2):
            if linha is None or not any(linha):
                continue
            nome, cpf, telefone, email, observacao = (list(linha) + [None] * 5)[:5]

            dados = {
                "nome": str(nome or "").strip(),
                "cpf": normalizar_cpf(str(cpf or "")),
                "telefone": normalizar_telefone(str(telefone or "")),
                "email": str(email or "").strip(),
                "observacao": str(observacao or "").strip(),
            }

            erro = validar_campos_cliente(dados)
            if erro:
                resultado["erros"].append({"linha": numero_linha, "erro": f"[Clientes] {erro}"})
                continue

            existente = (
                self._session.query(Cliente)
                .filter(Cliente.cpf == dados["cpf"])
                .first()
            )
            if existente:
                ok, erro = self._cliente_svc.atualizar(existente.id, dados)
                if not ok:
                    resultado["erros"].append({"linha": numero_linha, "erro": f"[Clientes] {erro}"})
                    continue
                cpf_para_cliente_id[dados["cpf"]] = existente.id
                resultado["clientes_atualizados"] += 1
            else:
                cliente, erro = self._cliente_svc.criar(dados)
                if not cliente:
                    resultado["erros"].append({"linha": numero_linha, "erro": f"[Clientes] {erro}"})
                    continue
                cpf_para_cliente_id[dados["cpf"]] = cliente.id
                resultado["clientes_criados"] += 1

    def _importar_veiculos(self, ws, resultado: dict, cpf_para_cliente_id: dict) -> None:
        linhas = ws.iter_rows(min_row=2, values_only=True)
        for numero_linha, linha in enumerate(linhas, start=2):
            if linha is None or not any(linha):
                continue
            cpf, placa, renavam, proprietario, marca_modelo, situacao, especie, observacao = (
                list(linha) + [None] * 8
            )[:8]

            cpf_normalizado = normalizar_cpf(str(cpf or ""))
            cliente_id = cpf_para_cliente_id.get(cpf_normalizado)
            if not cliente_id:
                cliente = (
                    self._session.query(Cliente)
                    .filter(Cliente.cpf == cpf_normalizado)
                    .first()
                )
                cliente_id = cliente.id if cliente else None

            if not cliente_id:
                resultado["erros"].append({
                    "linha": numero_linha,
                    "erro": f"[Veículos] Nenhum cliente encontrado com o CPF '{cpf}'. "
                            f"Cadastre o cliente antes (ou inclua-o na aba Clientes desta planilha).",
                })
                continue

            situacao = str(situacao or "ativo").strip().lower() or "ativo"
            especie = str(especie or "passeio").strip().lower() or "passeio"
            if situacao not in SITUACOES_VALIDAS:
                situacao = "ativo"
            if especie not in ESPECIES_VALIDAS:
                especie = "passeio"

            dados = {
                "cliente_id": cliente_id,
                "placa": str(placa or "").strip(),
                "renavam": str(renavam or "").strip(),
                "proprietario": str(proprietario or "").strip(),
                "marca_modelo": str(marca_modelo or "").strip(),
                "situacao": situacao,
                "especie": especie,
                "observacao": str(observacao or "").strip(),
            }

            erro = validar_campos_veiculo(dados)
            if erro:
                resultado["erros"].append({"linha": numero_linha, "erro": f"[Veículos] {erro}"})
                continue

            from app.models.veiculo import Veiculo
            existente = (
                self._session.query(Veiculo)
                .filter(Veiculo.cliente_id == cliente_id, Veiculo.placa == dados["placa"].upper())
                .first()
            )
            if existente:
                ok, erro = self._veiculo_svc.atualizar(existente.id, dados)
                if not ok:
                    resultado["erros"].append({"linha": numero_linha, "erro": f"[Veículos] {erro}"})
                    continue
                resultado["veiculos_atualizados"] += 1
            else:
                veiculo, erro = self._veiculo_svc.criar(dados)
                if not veiculo:
                    resultado["erros"].append({"linha": numero_linha, "erro": f"[Veículos] {erro}"})
                    continue
                resultado["veiculos_criados"] += 1
