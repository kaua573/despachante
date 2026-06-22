"""
Serviço de relatórios configuráveis.

Responsável por:
- Montar queries dinâmicas com base na configuração do usuário
- Gerar saída em PDF (ReportLab) ou Excel (openpyxl)
- Persistir e recuperar templates de relatório salvos
"""
from __future__ import annotations

import io
import json
from datetime import date, datetime
from typing import Any, Optional

from sqlalchemy.orm import Session

from app.models.cliente import Cliente
from app.models.ipva import Ipva
from app.models.licenciamento import Licenciamento
from app.models.multa import Multa
from app.models.template_relatorio import TemplateRelatorio
from app.models.veiculo import Veiculo


# ── Definição de campos disponíveis por domínio ───────────────────────────

CAMPOS_IPVA = [
    {"id": "cliente_nome",    "label": "Cliente"},
    {"id": "placa",           "label": "Placa"},
    {"id": "proprietario",    "label": "Proprietário"},
    {"id": "marca_modelo",    "label": "Marca/Modelo"},
    {"id": "especie",         "label": "Espécie"},
    {"id": "situacao",        "label": "Situação do Veículo"},
    {"id": "ano_referencia",  "label": "Ano Referência"},
    {"id": "valor",           "label": "Valor"},
    {"id": "vencimento",      "label": "Vencimento"},
    {"id": "status_pag",      "label": "Status"},
    {"id": "data_pagamento",  "label": "Data Pagamento"},
    {"id": "observacao",      "label": "Observação"},
]

CAMPOS_LICENCIAMENTO = [
    {"id": "cliente_nome",    "label": "Cliente"},
    {"id": "placa",           "label": "Placa"},
    {"id": "proprietario",    "label": "Proprietário"},
    {"id": "marca_modelo",    "label": "Marca/Modelo"},
    {"id": "especie",         "label": "Espécie"},
    {"id": "situacao",        "label": "Situação do Veículo"},
    {"id": "ano_referencia",  "label": "Ano Referência"},
    {"id": "valor",           "label": "Valor"},
    {"id": "vencimento",      "label": "Vencimento"},
    {"id": "status_pag",      "label": "Status"},
    {"id": "data_pagamento",  "label": "Data Pagamento"},
    {"id": "observacao",      "label": "Observação"},
]

CAMPOS_MULTAS = [
    {"id": "cliente_nome",    "label": "Cliente"},
    {"id": "placa",           "label": "Placa"},
    {"id": "proprietario",    "label": "Proprietário"},
    {"id": "marca_modelo",    "label": "Marca/Modelo"},
    {"id": "especie",         "label": "Espécie"},
    {"id": "auto_infracao",   "label": "Auto de Infração"},
    {"id": "data_infracao",   "label": "Data da Infração"},
    {"id": "descricao",       "label": "Descrição"},
    {"id": "valor",           "label": "Valor"},
    {"id": "vencimento",      "label": "Vencimento"},
    {"id": "status_pag",      "label": "Status"},
    {"id": "data_pagamento",  "label": "Data Pagamento"},
    {"id": "observacao",      "label": "Observação"},
]

CAMPOS_POR_TIPO = {
    "ipva": CAMPOS_IPVA,
    "licenciamento": CAMPOS_LICENCIAMENTO,
    "multas": CAMPOS_MULTAS,
}


class RelatorioService:
    def __init__(self, session: Session) -> None:
        self._session = session

    # ── Templates salvos ────────────────────────────────────────────────────

    def listar_templates(self) -> list[TemplateRelatorio]:
        return self._session.query(TemplateRelatorio).order_by(TemplateRelatorio.nome).all()

    def obter_template(self, template_id: int) -> Optional[TemplateRelatorio]:
        return self._session.get(TemplateRelatorio, template_id)

    def salvar_template(self, nome: str, config: dict, template_id: Optional[int] = None) -> TemplateRelatorio:
        config_json = json.dumps(config, ensure_ascii=False)
        if template_id:
            t = self._session.get(TemplateRelatorio, template_id)
            if t:
                t.nome = nome
                t.config_json = config_json
                t.atualizado_em = datetime.now()
                self._session.commit()
                return t
        t = TemplateRelatorio(nome=nome, config_json=config_json)
        self._session.add(t)
        self._session.commit()
        return t

    def excluir_template(self, template_id: int) -> tuple[bool, str]:
        t = self._session.get(TemplateRelatorio, template_id)
        if not t:
            return False, "Template não encontrado."
        self._session.delete(t)
        self._session.commit()
        return True, ""

    # ── Busca de dados ──────────────────────────────────────────────────────

    def buscar_dados(self, config: dict) -> list[dict]:
        """
        Executa a query de acordo com a configuração passada.
        Retorna lista de dicionários com os dados brutos (todos os campos),
        sem filtrar por campos visíveis — isso é feito na renderização.
        """
        tipo = config.get("tipo", "ipva")
        filtros = config.get("filtros", {})
        ordenar_por = config.get("ordenar_por", "vencimento")
        ordem_direcao = config.get("ordem_direcao", "asc")

        if tipo == "ipva":
            return self._buscar_ipva(filtros, ordenar_por, ordem_direcao)
        elif tipo == "licenciamento":
            return self._buscar_licenciamento(filtros, ordenar_por, ordem_direcao)
        elif tipo == "multas":
            return self._buscar_multas(filtros, ordenar_por, ordem_direcao)
        return []

    def _buscar_ipva(self, filtros: dict, ordenar_por: str, direcao: str) -> list[dict]:
        q = (
            self._session.query(Ipva, Veiculo, Cliente)
            .join(Veiculo, Ipva.veiculo_id == Veiculo.id)
            .join(Cliente, Veiculo.cliente_id == Cliente.id)
        )
        q = self._aplicar_filtros_comuns(q, Ipva, filtros)
        q = self._aplicar_ordenacao(q, Ipva, Veiculo, Cliente, ordenar_por, direcao)
        return [self._montar_linha_ipva(r, v, c) for r, v, c in q.all()]

    def _buscar_licenciamento(self, filtros: dict, ordenar_por: str, direcao: str) -> list[dict]:
        q = (
            self._session.query(Licenciamento, Veiculo, Cliente)
            .join(Veiculo, Licenciamento.veiculo_id == Veiculo.id)
            .join(Cliente, Veiculo.cliente_id == Cliente.id)
        )
        q = self._aplicar_filtros_comuns(q, Licenciamento, filtros)
        q = self._aplicar_ordenacao(q, Licenciamento, Veiculo, Cliente, ordenar_por, direcao)
        return [self._montar_linha_licenciamento(r, v, c) for r, v, c in q.all()]

    def _buscar_multas(self, filtros: dict, ordenar_por: str, direcao: str) -> list[dict]:
        q = (
            self._session.query(Multa, Veiculo, Cliente)
            .join(Veiculo, Multa.veiculo_id == Veiculo.id)
            .join(Cliente, Veiculo.cliente_id == Cliente.id)
        )
        q = self._aplicar_filtros_multas(q, filtros)
        q = self._aplicar_ordenacao(q, Multa, Veiculo, Cliente, ordenar_por, direcao)
        return [self._montar_linha_multa(r, v, c) for r, v, c in q.all()]

    def _aplicar_filtros_comuns(self, q, modelo, filtros: dict):
        """Aplica filtros de data, status, placa e cliente — comuns a IPVA e licenciamento."""
        if filtros.get("data_inicio"):
            q = q.filter(modelo.vencimento >= filtros["data_inicio"])
        if filtros.get("data_fim"):
            q = q.filter(modelo.vencimento <= filtros["data_fim"])
        if filtros.get("status") == "pago":
            q = q.filter(modelo.pago == True)
        elif filtros.get("status") in ("pendente", "vencido"):
            q = q.filter(modelo.pago == False)
            if filtros["status"] == "vencido":
                q = q.filter(modelo.vencimento < date.today().isoformat())
        if filtros.get("placa"):
            q = q.filter(Veiculo.placa.ilike(f"%{filtros['placa']}%"))
        if filtros.get("cliente_nome"):
            q = q.filter(Cliente.nome.ilike(f"%{filtros['cliente_nome']}%"))
        return q

    def _aplicar_filtros_multas(self, q, filtros: dict):
        if filtros.get("data_inicio"):
            q = q.filter(Multa.data_infracao >= filtros["data_inicio"])
        if filtros.get("data_fim"):
            q = q.filter(Multa.data_infracao <= filtros["data_fim"])
        if filtros.get("status") == "pago":
            q = q.filter(Multa.pago == True)
        elif filtros.get("status") in ("pendente", "vencido"):
            q = q.filter(Multa.pago == False)
            if filtros["status"] == "vencido":
                q = q.filter(Multa.vencimento < date.today().isoformat())
        if filtros.get("placa"):
            q = q.filter(Veiculo.placa.ilike(f"%{filtros['placa']}%"))
        if filtros.get("cliente_nome"):
            q = q.filter(Cliente.nome.ilike(f"%{filtros['cliente_nome']}%"))
        return q

    def _aplicar_ordenacao(self, q, modelo, veiculo_model, cliente_model, campo: str, direcao: str):
        """Mapeia nome de campo para coluna SQLAlchemy e aplica ordenação."""
        mapa = {
            "cliente_nome":   cliente_model.nome,
            "placa":          veiculo_model.placa,
            "proprietario":   veiculo_model.proprietario,
            "vencimento":     modelo.vencimento,
            "valor":          modelo.valor,
            "status_pag":     modelo.pago,
        }
        # Campos específicos por tipo
        if hasattr(modelo, "ano_referencia"):
            mapa["ano_referencia"] = modelo.ano_referencia
        if hasattr(modelo, "data_infracao"):
            mapa["data_infracao"] = modelo.data_infracao

        coluna = mapa.get(campo, modelo.vencimento if hasattr(modelo, "vencimento") else modelo.id)
        if direcao == "desc":
            coluna = coluna.desc()
        return q.order_by(coluna)

    # ── Montagem de linhas ──────────────────────────────────────────────────

    def _montar_linha_ipva(self, r: Ipva, v: Veiculo, c: Cliente) -> dict:
        return {
            "cliente_nome":   c.nome,
            "placa":          v.placa,
            "proprietario":   v.proprietario or "",
            "marca_modelo":   v.marca_modelo or "",
            "especie":        v.especie or "",
            "situacao":       v.situacao or "",
            "ano_referencia": r.ano_referencia,
            "valor":          r.valor,
            "vencimento":     r.vencimento or "",
            "status_pag":     "Pago" if r.pago else self._status_vencimento(r.vencimento),
            "data_pagamento": r.data_pagamento or "",
            "observacao":     r.observacao or "",
        }

    def _montar_linha_licenciamento(self, r: Licenciamento, v: Veiculo, c: Cliente) -> dict:
        return {
            "cliente_nome":   c.nome,
            "placa":          v.placa,
            "proprietario":   v.proprietario or "",
            "marca_modelo":   v.marca_modelo or "",
            "especie":        v.especie or "",
            "situacao":       v.situacao or "",
            "ano_referencia": r.ano_referencia,
            "valor":          r.valor,
            "vencimento":     r.vencimento or "",
            "status_pag":     "Pago" if r.pago else self._status_vencimento(r.vencimento),
            "data_pagamento": r.data_pagamento or "",
            "observacao":     r.observacao or "",
        }

    def _montar_linha_multa(self, r: Multa, v: Veiculo, c: Cliente) -> dict:
        return {
            "cliente_nome":   c.nome,
            "placa":          v.placa,
            "proprietario":   v.proprietario or "",
            "marca_modelo":   v.marca_modelo or "",
            "especie":        v.especie or "",
            "auto_infracao":  r.auto_infracao or "",
            "data_infracao":  r.data_infracao or "",
            "descricao":      r.descricao or "",
            "valor":          r.valor,
            "vencimento":     r.vencimento or "",
            "status_pag":     "Pago" if r.pago else self._status_vencimento(r.vencimento),
            "data_pagamento": r.data_pagamento or "",
            "observacao":     r.observacao or "",
        }

    @staticmethod
    def _status_vencimento(vencimento: Optional[str]) -> str:
        if not vencimento:
            return "Pendente"
        return "Vencido" if vencimento < date.today().isoformat() else "Pendente"

    # ── Agrupamento ─────────────────────────────────────────────────────────

    @staticmethod
    def agrupar(dados: list[dict], campo_grupo: str) -> dict[str, list[dict]]:
        """Agrupa lista de dicts por valor de um campo. Retorna OrderedDict implícito."""
        grupos: dict[str, list[dict]] = {}
        for linha in dados:
            chave = str(linha.get(campo_grupo, "—") or "—")
            grupos.setdefault(chave, []).append(linha)
        return grupos

    # ── Totalizadores ───────────────────────────────────────────────────────

    @staticmethod
    def calcular_totais(dados: list[dict]) -> dict:
        total_valor = sum(float(r.get("valor") or 0) for r in dados)
        total_registros = len(dados)
        total_pago = sum(1 for r in dados if r.get("status_pag") == "Pago")
        total_pendente = sum(1 for r in dados if r.get("status_pag") == "Pendente")
        total_vencido = sum(1 for r in dados if r.get("status_pag") == "Vencido")
        valor_pago = sum(float(r.get("valor") or 0) for r in dados if r.get("status_pag") == "Pago")
        valor_pendente = sum(float(r.get("valor") or 0) for r in dados if r.get("status_pag") != "Pago")
        return {
            "total_registros": total_registros,
            "total_valor": total_valor,
            "total_pago": total_pago,
            "total_pendente": total_pendente,
            "total_vencido": total_vencido,
            "valor_pago": valor_pago,
            "valor_pendente": valor_pendente,
        }

    # ── Exportação PDF ──────────────────────────────────────────────────────

    def gerar_pdf(
        self,
        dados: list[dict],
        campos_visiveis: list[dict],
        config: dict,
        cfg_pdf: dict,
    ) -> bytes:
        """
        Gera PDF do relatório.

        Args:
            dados: linhas do relatório
            campos_visiveis: lista de {id, label} na ordem desejada
            config: configuração do relatório (tipo, filtros, agrupamento etc.)
            cfg_pdf: configurações visuais do PDF (fonte, cor, etc.)
        """
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as rl_colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, HRFlowable, KeepTogether,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from app.services.configuracao_service import FONTES_PDF, TAMANHOS_PDF, PALETA_CORES

        # Resolve configurações visuais
        fonte = FONTES_PDF.get(cfg_pdf.get("fonte", "moderna"), FONTES_PDF["moderna"])
        tam = TAMANHOS_PDF.get(cfg_pdf.get("tamanho", "medio"), TAMANHOS_PDF["medio"])
        cor_chave = cfg_pdf.get("cor", "azul")
        paleta = PALETA_CORES.get(cor_chave, PALETA_CORES["azul"])
        mostrar_data = cfg_pdf.get("mostrar_data_geracao", True)
        nome_escritorio = cfg_pdf.get("nome_escritorio", "")

        COR_PRINCIPAL = rl_colors.HexColor(paleta["principal"])
        COR_SECUND = rl_colors.HexColor(paleta["secundaria"])
        CINZA1 = rl_colors.HexColor("#f4f6fa")
        CINZA2 = rl_colors.HexColor("#e8ecf2")
        BRANCO = rl_colors.white

        FONTE_BASE = fonte["base"]
        FONTE_BOLD = fonte["bold"]

        sTitulo = ParagraphStyle("titulo", fontName=FONTE_BOLD, fontSize=tam["titulo"], textColor=COR_PRINCIPAL, spaceAfter=4)
        sSub    = ParagraphStyle("sub",    fontName=FONTE_BASE, fontSize=tam["mini"] + 1, textColor=rl_colors.HexColor("#5a6680"), spaceAfter=10)
        sSecao  = ParagraphStyle("secao",  fontName=FONTE_BOLD, fontSize=tam["secao"],  textColor=COR_SECUND, spaceBefore=12, spaceAfter=4)
        sCell   = ParagraphStyle("cell",   fontName=FONTE_BASE, fontSize=tam["texto"] - 1, textColor=rl_colors.HexColor("#1c2333"))
        sHead   = ParagraphStyle("head",   fontName=FONTE_BOLD, fontSize=tam["mini"],   textColor=BRANCO)
        sTotal  = ParagraphStyle("total",  fontName=FONTE_BOLD, fontSize=tam["texto"] - 1, textColor=COR_PRINCIPAL)
        sRodape = ParagraphStyle("rodape", fontName=FONTE_BASE, fontSize=tam["mini"],   textColor=rl_colors.HexColor("#9aa4ba"))

        # Usa landscape se houver muitas colunas
        pagesize = landscape(A4) if len(campos_visiveis) > 6 else A4

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=pagesize,
            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
        )

        story = []

        # Cabeçalho
        if nome_escritorio:
            story.append(Paragraph(nome_escritorio, ParagraphStyle(
                "escrit", fontName=FONTE_BOLD, fontSize=tam["mini"] + 3, textColor=COR_SECUND, spaceAfter=2
            )))
        tipo_label = {"ipva": "IPVA", "licenciamento": "Licenciamento", "multas": "Multas"}.get(config.get("tipo", "ipva"), "Relatório")
        story.append(Paragraph(f"Relatório — {tipo_label}", sTitulo))
        if mostrar_data:
            story.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", sSub))
        story.append(HRFlowable(width="100%", thickness=2, color=COR_PRINCIPAL, spaceAfter=12))

        # Resumo executivo (opcional)
        if config.get("mostrar_resumo"):
            totais = self.calcular_totais(dados)
            story.append(Paragraph("Resumo Executivo", sSecao))
            resumo_data = [
                [Paragraph("Total de registros", sHead), Paragraph(str(totais["total_registros"]), sTotal)],
                [Paragraph("Valor total", sHead), Paragraph(self._fmt_moeda(totais["total_valor"]), sTotal)],
                [Paragraph("Pagos", sHead), Paragraph(f"{totais['total_pago']} ({self._fmt_moeda(totais['valor_pago'])})", sTotal)],
                [Paragraph("Pendentes", sHead), Paragraph(f"{totais['total_pendente']}", sTotal)],
                [Paragraph("Vencidos", sHead), Paragraph(f"{totais['total_vencido']} ({self._fmt_moeda(totais['valor_pendente'])})", sTotal)],
            ]
            t_resumo = Table(resumo_data, colWidths=["40%", "60%"])
            t_resumo.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (0, -1), COR_PRINCIPAL),
                ("BACKGROUND", (1, 0), (1, -1), CINZA1),
                ("ROWBACKGROUNDS", (1, 0), (1, -1), [CINZA1, CINZA2]),
                ("BOX", (0, 0), (-1, -1), 0.5, CINZA2),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, CINZA2),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]))
            story.append(t_resumo)
            story.append(Spacer(1, 14))

        # Tabela de dados (com ou sem agrupamento)
        campo_grupo = config.get("agrupar_por", "")
        if campo_grupo and campo_grupo != "nenhum":
            grupos = self.agrupar(dados, campo_grupo)
            for nome_grupo, linhas in grupos.items():
                story.append(Paragraph(str(nome_grupo), sSecao))
                story.extend(self._montar_tabela_pdf(linhas, campos_visiveis, sHead, sCell, COR_PRINCIPAL, CINZA1, CINZA2, BRANCO, config))
                story.append(Spacer(1, 8))
        else:
            story.extend(self._montar_tabela_pdf(dados, campos_visiveis, sHead, sCell, COR_PRINCIPAL, CINZA1, CINZA2, BRANCO, config))

        # Rodapé
        story.append(Spacer(1, 16))
        story.append(HRFlowable(width="100%", thickness=1, color=CINZA2))
        story.append(Paragraph(nome_escritorio or "Sistema de Despachante", sRodape))

        doc.build(story)
        buf.seek(0)
        return buf.read()

    def _montar_tabela_pdf(self, dados, campos_visiveis, sHead, sCell, cor_principal, cinza1, cinza2, branco, config):
        from reportlab.platypus import Table, TableStyle, Spacer, Paragraph

        if not dados:
            return [Paragraph("Nenhum registro encontrado.", sCell)]

        cabecalho = [Paragraph(c["label"], sHead) for c in campos_visiveis]
        linhas = [cabecalho]
        for linha in dados:
            row = []
            for campo in campos_visiveis:
                val = linha.get(campo["id"], "")
                if campo["id"] == "valor" and val is not None:
                    val = self._fmt_moeda(val)
                elif campo["id"] in ("vencimento", "data_pagamento", "data_infracao"):
                    val = self._fmt_data(str(val))
                row.append(Paragraph(str(val) if val is not None else "—", sCell))
            linhas.append(row)

        # Linha de totais
        if config.get("mostrar_totais") and dados:
            totais = self.calcular_totais(dados)
            linha_total = ["—"] * len(campos_visiveis)
            ids = [c["id"] for c in campos_visiveis]
            if "valor" in ids:
                linha_total[ids.index("valor")] = self._fmt_moeda(totais["total_valor"])
            if "cliente_nome" in ids:
                linha_total[ids.index("cliente_nome")] = f"Total: {totais['total_registros']} registros"
            linhas.append([Paragraph(v, sHead) for v in linha_total])

        # Largura proporcional das colunas
        n = len(campos_visiveis)
        col_w = [f"{100 / n:.1f}%"] * n

        t = Table(linhas, colWidths=col_w, repeatRows=1)
        estilos = [
            ("BACKGROUND", (0, 0), (-1, 0), cor_principal),
            ("TEXTCOLOR", (0, 0), (-1, 0), branco),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [branco, cinza1]),
            ("BOX", (0, 0), (-1, -1), 0.5, cinza2),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, cinza2),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]
        if config.get("mostrar_totais") and dados:
            # Destaca última linha como totais
            estilos.append(("BACKGROUND", (0, -1), (-1, -1), cor_principal))
        t.setStyle(TableStyle(estilos))
        return [t]

    # ── Exportação Excel ────────────────────────────────────────────────────

    def gerar_excel(self, dados: list[dict], campos_visiveis: list[dict], config: dict) -> bytes:
        """Gera planilha Excel com os dados do relatório."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        tipo_label = {"ipva": "IPVA", "licenciamento": "Licenciamento", "multas": "Multas"}.get(config.get("tipo", "ipva"), "Relatório")
        ws.title = tipo_label

        # Estilo do cabeçalho
        fill_header = PatternFill(start_color="1A4F8A", end_color="1A4F8A", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        fill_alt = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")
        borda = Border(
            left=Side(style="thin", color="C8D0DC"),
            right=Side(style="thin", color="C8D0DC"),
            top=Side(style="thin", color="C8D0DC"),
            bottom=Side(style="thin", color="C8D0DC"),
        )

        # Cabeçalho
        for col_idx, campo in enumerate(campos_visiveis, start=1):
            cell = ws.cell(row=1, column=col_idx, value=campo["label"])
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

        # Dados
        for row_idx, linha in enumerate(dados, start=2):
            fill_row = fill_alt if row_idx % 2 == 0 else None
            for col_idx, campo in enumerate(campos_visiveis, start=1):
                val = linha.get(campo["id"], "")
                if campo["id"] == "valor" and val is not None:
                    try:
                        val = float(val)
                    except (TypeError, ValueError):
                        val = ""
                elif campo["id"] in ("vencimento", "data_pagamento", "data_infracao"):
                    val = self._fmt_data(str(val)) if val else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
                if fill_row:
                    cell.fill = fill_row
                cell.border = borda
                cell.alignment = Alignment(vertical="center")

        # Linha de totais
        if config.get("mostrar_totais") and dados:
            totais = self.calcular_totais(dados)
            linha_total_idx = len(dados) + 2
            ids = [c["id"] for c in campos_visiveis]
            fill_total = PatternFill(start_color="1A4F8A", end_color="1A4F8A", fill_type="solid")
            for col_idx, campo in enumerate(campos_visiveis, start=1):
                val = ""
                if campo["id"] == "valor":
                    val = totais["total_valor"]
                elif campo["id"] == "cliente_nome":
                    val = f"TOTAL: {totais['total_registros']} registros"
                cell = ws.cell(row=linha_total_idx, column=col_idx, value=val)
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = fill_total
                cell.border = borda
                cell.alignment = Alignment(vertical="center")

        # Ajusta largura das colunas
        for col_idx, campo in enumerate(campos_visiveis, start=1):
            max_len = max(
                len(str(campo["label"])),
                *(len(str(linha.get(campo["id"], "") or "")) for linha in dados[:50]),
                10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── Helpers de formatação ───────────────────────────────────────────────

    @staticmethod
    def _fmt_data(valor: str) -> str:
        if not valor or len(valor) < 8:
            return "—"
        try:
            y, m, d = valor[:10].split("-")
            return f"{d}/{m}/{y}"
        except (ValueError, AttributeError):
            return valor

    @staticmethod
    def _fmt_moeda(valor) -> str:
        if valor is None or valor == "":
            return "—"
        try:
            return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (TypeError, ValueError):
            return "—"
